from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Count, Min as models_Min
from .models import InventoryItem, CATEGORY_CHOICES, CLASSIFICATION_CHOICES
from .forms import InventoryItemForm
import json


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'inventory/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def dashboard(request):
    import json as json_lib
    from django.db.models import Count

    categories = [c[0] for c in CATEGORY_CHOICES]
    category_counts = {}
    for cat in categories:
        category_counts[cat] = InventoryItem.objects.filter(category=cat).count()

    total = InventoryItem.objects.count()
    active = InventoryItem.objects.filter(status='Active').count()
    inactive = InventoryItem.objects.filter(status='Inactive').count()
    under_repair = InventoryItem.objects.filter(status='Under Repair').count()
    for_repair = InventoryItem.objects.filter(status='For Repair').count()
    disposed = InventoryItem.objects.filter(status='Disposed').count()

    health_score = round((active / total * 100) if total > 0 else 0)

    dept_data = (
        InventoryItem.objects.exclude(department='')
        .values('department')
        .annotate(count=Count('department'))
        .order_by('-count')[:8]
    )
    dept_labels = json_lib.dumps([d['department'] for d in dept_data])
    dept_counts_json = json_lib.dumps([d['count'] for d in dept_data])

    status_data = json_lib.dumps([active, inactive, under_repair, for_repair, disposed])
    cat_labels = json_lib.dumps([c for c in categories])
    cat_data = json_lib.dumps([category_counts.get(c, 0) for c in categories])

    site_data = (
        InventoryItem.objects.exclude(project_site='')
        .values('project_site')
        .annotate(count=Count('project_site'))
        .order_by('-count')[:6]
    )

    insights = []
    if total > 0:
        repair_pct = round((under_repair + for_repair) / total * 100)
        inactive_pct = round(inactive / total * 100)
        if repair_pct >= 10:
            insights.append({'level': 'danger', 'icon': 'bi-tools', 'title': 'High Repair Rate',
                'msg': f'{repair_pct}% of assets ({under_repair + for_repair} items) need repair attention.'})
        if inactive_pct >= 15:
            insights.append({'level': 'warning', 'icon': 'bi-moon-stars', 'title': 'Idle Assets Detected',
                'msg': f'{inactive_pct}% ({inactive} items) are inactive — consider redeployment.'})
        top_cat = max(category_counts, key=category_counts.get) if category_counts else None
        if top_cat and category_counts[top_cat] > 0:
            insights.append({'level': 'info', 'icon': 'bi-bar-chart', 'title': f'{top_cat} Dominates',
                'msg': f'{top_cat} is the largest category with {category_counts[top_cat]} assets ({round(category_counts[top_cat]/total*100)}% of total).'})
        if health_score >= 85:
            insights.append({'level': 'success', 'icon': 'bi-shield-check', 'title': 'Fleet in Good Health',
                'msg': f'{health_score}% of assets are active — your inventory is well-maintained.'})
        if disposed > 0:
            insights.append({'level': 'secondary', 'icon': 'bi-trash3', 'title': 'Disposed Assets',
                'msg': f'{disposed} asset{"s" if disposed > 1 else ""} marked as disposed. Consider archiving records.'})

    # Classification counts for new chart
    class_data = (
        InventoryItem.objects.values('classification')
        .annotate(count=Count('classification'))
        .order_by('-count')
    )
    class_labels = json_lib.dumps([d['classification'] for d in class_data])
    class_counts  = json_lib.dumps([d['count'] for d in class_data])

    context = {
        'categories': categories,
        'category_counts': category_counts,
        'total': total,
        'active': active,
        'inactive': inactive,
        'under_repair': under_repair,
        'for_repair': for_repair,
        'disposed': disposed,
        'health_score': health_score,
        'dept_labels': dept_labels,
        'dept_counts_json': dept_counts_json,
        'status_data': status_data,
        'cat_labels': cat_labels,
        'cat_data': cat_data,
        'site_data': site_data,
        'insights': insights,
        'class_labels': class_labels,
        'class_counts': class_counts,
        'recent_items': InventoryItem.objects.all()[:10],
    }
    return render(request, 'inventory/dashboard.html', context)


@login_required
def inventory_list(request, category=None, classification=None):
    items = InventoryItem.objects.all()

    if category:
        items = items.filter(category=category)
    if classification:
        items = items.filter(classification=classification)

    # Search
    search = request.GET.get('search', '')
    if search:
        items = items.filter(
            Q(id_number__icontains=search) |
            Q(asset_tag__icontains=search) |
            Q(brand__icontains=search) |
            Q(model__icontains=search) |
            Q(assigned_person__icontains=search) |
            Q(department__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(reference_number__icontains=search)
        )

    # Filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        items = items.filter(status=status_filter)

    dept_filter = request.GET.get('department', '')
    if dept_filter:
        items = items.filter(department__icontains=dept_filter)

    context = {
        'items': items,
        'category': category,
        'classification': classification,
        'categories': [c[0] for c in CATEGORY_CHOICES],
        'classifications': [c[0] for c in CLASSIFICATION_CHOICES],
        'search': search,
        'status_filter': status_filter,
        'dept_filter': dept_filter,
        'form': InventoryItemForm(),
    }
    return render(request, 'inventory/inventory_list.html', context)


@login_required
def add_item(request):
    if request.method == 'POST':
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Item added successfully.')
        else:
            messages.error(request, 'Error adding item: ' + str(form.errors))
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


@login_required
def edit_item(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            messages.success(request, 'Item updated successfully.')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
            messages.error(request, 'Error updating item.')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


@login_required
def get_item(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    data = {
        'id': item.pk,
        'id_number': item.id_number,
        'asset_tag': item.asset_tag,
        'computer_tag': item.computer_tag,
        'display_tag': item.display_tag,
        'brand': item.brand,
        'model': item.model,
        'work_order': item.work_order,
        'inventory_receipt': item.inventory_receipt,
        'inventory_issuance': item.inventory_issuance,
        'po': item.po,
        'remarks': item.remarks,
        'assigned_person': item.assigned_person,
        'department': item.department,
        'project_site': item.project_site,
        'serial_number': item.serial_number,
        'type': item.type,
        'status': item.status,
        'reference_number': item.reference_number,
        'purchase_date': item.purchase_date.isoformat() if item.purchase_date else '',
        'category': item.category,
        'classification': item.classification,
    }
    return JsonResponse(data)


@login_required
def delete_item(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    item.delete()
    messages.success(request, 'Item deleted.')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


# ─── Barcode Views ───────────────────────────────────────────

from .barcode_utils import generate_barcode_base64, save_barcode_file


# ─── Excel Upload Views ───────────────────────────────────────

import openpyxl
from django.http import HttpResponse
import io

EXCEL_COLUMNS = [
    'id_number', 'asset_tag', 'computer_tag', 'display_tag',
    'brand', 'model', 'category', 'classification', 'type',
    'status', 'serial_number', 'assigned_person', 'department',
    'project_site', 'purchase_date', 'po', 'work_order',
    'inventory_receipt', 'inventory_issuance', 'reference_number', 'remarks',
]

CATEGORY_VALUES = [c[0] for c in CATEGORY_CHOICES]
CLASSIFICATION_VALUES = [c[0] for c in CLASSIFICATION_CHOICES]
STATUS_VALUES = ['Active', 'Inactive', 'Under Repair', 'For Repair', 'Disposed']


@login_required
def download_excel_template(request):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    COLUMNS = [
        ('ID Number',                  14),
        ('Asset Tag',                  14),
        ('Computer Tag',               14),
        ('Display Tag',                13),
        ('Brand',                      14),
        ('Model',                      20),
        ('Category *',                 18),
        ('Classification *',           16),
        ('Type',                       14),
        ('Status *',                   13),
        ('Serial Number',              18),
        ('Assigned Person',            22),
        ('Department',                 20),
        ('Project/Site',               18),
        ('Purchase Date (YYYY-MM-DD)', 24),
        ('PO',                         12),
        ('Work Order',                 14),
        ('Inventory Receipt',          18),
        ('Inventory Issuance',         18),
        ('Reference Number',           18),
        ('Remarks',                    30),
    ]
    CAT_COLORS = {
        'Desktop': 'DBEAFE', 'Laptop': 'D1FAE5', 'Printer': 'FEF3C7',
        'Security': 'EDE9FE', 'Network Devices': 'E0F2FE',
        'Electrical Devices': 'FEF9C3', 'Tools': 'F3F4F6',
        'Telephony': 'FCE7F3', 'Display': 'CCFBF1',
    }
    SAMPLES = [
        ('IT-001','AT-DT-001','CT-001','','Dell','OptiPlex 7090','Desktop','Components','Desktop PC','Active','SN001','Juan Dela Cruz','IT Dept','Main Office','2024-01-15','','','','','',''),
        ('IT-002','AT-LT-001','','','Lenovo','ThinkPad X1 Carbon','Laptop','Components','Laptop','Active','SN002','Maria Santos','Finance','Main Office','2024-02-10','','','','','',''),
        ('IT-003','AT-PR-001','','','HP','LaserJet Pro M404n','Printer','Parts','Laser Printer','Active','SN003','Jose Reyes','Admin','Branch A','2023-11-20','','','','','',''),
        ('IT-004','AT-SC-001','','','Hikvision','DS-2CD2143G2-I','Security','Components','IP Camera','Active','SN004','Pedro Cruz','Security','Main Gate','2023-06-01','','','','','',''),
        ('IT-005','AT-NW-001','','','Cisco','SG350-28','Network Devices','Components','Managed Switch','Active','SN005','','IT Dept','Server Room','2022-09-15','','','','','',''),
        ('IT-006','AT-ED-001','','','APC','Smart-UPS 1500','Electrical Devices','Components','UPS','Under Repair','SN006','','IT Dept','Server Room','2021-03-10','','','','','',''),
        ('IT-007','AT-TL-001','','','Generic','Crimping Kit','Tools','Parts','Network Tool','Active','SN007','','IT Dept','Main Office','','','','','','',''),
        ('IT-008','AT-TP-001','','','Cisco','CP-7841','Telephony','Components','IP Phone','Inactive','SN008','Ana Reyes','Reception','Main Office','2020-07-22','','','','','',''),
        ('IT-009','AT-DP-001','','DT-001','Samsung','S27A600UUE','Display','Components','Monitor','Active','SN009','Luis Garcia','IT Dept','Main Office','2023-12-05','','','','','',''),
    ]

    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('All Items')

    # Header
    for col_idx, (label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(bold=True, color='FFFFFF', name='Segoe UI', size=10)
        cell.fill      = PatternFill('solid', start_color='1A2B4A')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 38

    # Dropdowns
    cat_dv = DataValidation(type='list', formula1='"' + ','.join(CATEGORY_VALUES) + '"', showDropDown=False, allow_blank=False)
    cat_dv.sqref = 'G2:G50000'
    ws.add_data_validation(cat_dv)
    clf_dv = DataValidation(type='list', formula1='"' + ','.join(CLASSIFICATION_VALUES) + '"', showDropDown=False, allow_blank=False)
    clf_dv.sqref = 'H2:H50000'
    ws.add_data_validation(clf_dv)
    status_dv = DataValidation(type='list', formula1='"' + ','.join(STATUS_VALUES) + '"', showDropDown=False, allow_blank=True)
    status_dv.sqref = 'J2:J50000'
    ws.add_data_validation(status_dv)

    # Sample rows
    for row_idx, sample in enumerate(SAMPLES, 2):
        color = CAT_COLORS.get(sample[6], 'FFFFFF')
        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else None)
            cell.font      = Font(name='Segoe UI', size=9, italic=True, color='6B7280')
            cell.fill      = PatternFill('solid', start_color=color)
            cell.border    = border
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}1'

    # Instructions sheet
    wi = wb.create_sheet('Instructions')
    lines = [
        ('INVENTORY UPLOAD TEMPLATE — SINGLE SHEET', True,  '1A2B4A', 'FFFFFF', 13),
        ('', False, None, None, 10),
        ('HOW TO USE', True, 'EFF6FF', '1E40AF', 11),
        ('1. Go to the "All Items" sheet', False, None, None, 10),
        ('2. Delete the sample rows (rows 2–10) — they are examples only', False, None, None, 10),
        ('3. Enter ALL your inventory items here in any order — mix all categories freely', False, None, None, 10),
        ('4. Use the dropdowns for Category *, Classification *, and Status columns', False, None, None, 10),
        ('5. Save, then upload via the Upload Excel button — the system auto-sorts everything', False, None, None, 10),
        ('', False, None, None, 10),
        ('REQUIRED FIELDS', True, 'FEF3C7', '92400E', 11),
        ('  Category *        →  ' + ' / '.join(CATEGORY_VALUES), False, None, None, 10),
        ('  Classification *  →  Components  or  Parts', False, None, None, 10),
        ('  Status            →  ' + ' / '.join(STATUS_VALUES) + '  (defaults to Active if blank)', False, None, None, 10),
        ('', False, None, None, 10),
        ('DATE FORMAT', True, 'F0FDF4', '166534', 11),
        ('  Purchase Date must use YYYY-MM-DD format  (example: 2024-01-15)', False, None, None, 10),
        ('', False, None, None, 10),
        ('TIPS', True, 'F5F3FF', '5B21B6', 11),
        ('  • One sheet for everything — no switching tabs per category', False, None, None, 10),
        ('  • Blank rows are skipped automatically', False, None, None, 10),
        ('  • All fields except Category and Classification are optional', False, None, None, 10),
    ]
    for row_idx, (text, bold, bg, fg, size) in enumerate(lines, 1):
        cell = wi.cell(row=row_idx, column=1, value=text)
        cell.font      = Font(bold=bold, name='Segoe UI', size=size, color=fg or '374151')
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)
    wi.column_dimensions['A'].width = 95

    wb.active = wb['All Items']
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventory_upload_template.xlsx"'
    return response


@login_required
def upload_excel(request):
    if request.method != 'POST':
        return redirect('inventory_list')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'No file selected.')
        return redirect('inventory_list')

    if not excel_file.name.endswith(('.xlsx', '.xlsm')):
        messages.error(request, 'Invalid file type. Please upload an .xlsx file.')
        return redirect('inventory_list')

    try:
        from datetime import datetime, date as date_type

        wb = openpyxl.load_workbook(excel_file, data_only=True)

        TEMPLATE_LABELS = {
            'ID Number': 'id_number', 'Asset Tag': 'asset_tag',
            'Computer Tag': 'computer_tag', 'Display Tag': 'display_tag',
            'Brand': 'brand', 'Model': 'model', 'Category *': 'category',
            'Classification *': 'classification', 'Type': 'type',
            'Status *': 'status', 'Serial Number': 'serial_number',
            'Assigned Person': 'assigned_person', 'Department': 'department',
            'Project/Site': 'project_site',
            'Purchase Date (YYYY-MM-DD)': 'purchase_date',
            'PO': 'po', 'Work Order': 'work_order',
            'Inventory Receipt': 'inventory_receipt',
            'Inventory Issuance': 'inventory_issuance',
            'Reference Number': 'reference_number', 'Remarks': 'remarks',
        }

        # Detect format:
        # 1. Single "All Items" sheet (new unified template)
        # 2. Per-category sheets (Desktop, Laptop, etc.)
        # 3. Single "Inventory" sheet (legacy)
        # 4. Active sheet fallback
        category_sheets = [s for s in wb.sheetnames if s in CATEGORY_VALUES]
        has_all_items = 'All Items' in wb.sheetnames
        has_inventory_sheet = 'Inventory' in wb.sheetnames

        if has_all_items:
            sheets_to_process = [(wb['All Items'], None)]
        elif category_sheets:
            sheets_to_process = [(wb[s], s) for s in category_sheets]
        elif has_inventory_sheet:
            sheets_to_process = [(wb['Inventory'], None)]
        else:
            sheets_to_process = [(wb.active, None)]

        def build_col_map(ws):
            header_raw = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            col_map = {}
            for i, h in enumerate(header_raw):
                field = TEMPLATE_LABELS.get(h)
                if field:
                    col_map[field] = i
            return col_map

        def parse_date(val, raw_row, col_map):
            if val is None or str(val).strip() == '':
                return None
            idx = col_map.get('purchase_date', 0)
            raw = raw_row[idx] if idx < len(raw_row) else None
            if isinstance(raw, (datetime, date_type)):
                return raw.date() if hasattr(raw, 'date') else raw
            try:
                return datetime.strptime(str(val).strip()[:10], '%Y-%m-%d').date()
            except ValueError:
                return None

        def import_sheet(ws, forced_category=None, batch_id=''):
            col_map = build_col_map(ws)
            if 'category' not in col_map and not forced_category:
                return 0, 0, [f'Sheet "{ws.title}": Could not find required columns — skipped.']

            sheet_created, sheet_skipped, sheet_errors = 0, 0, []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(v is None or str(v).strip() == '' for v in row):
                    continue

                def get(field):
                    idx = col_map.get(field)
                    if idx is None or idx >= len(row):
                        return ''
                    val = row[idx]
                    return str(val).strip() if val is not None else ''

                category = forced_category or get('category')
                if category not in CATEGORY_VALUES:
                    sheet_errors.append(f'Sheet "{ws.title}" Row {row_num}: Invalid category "{category}" — skipped.')
                    sheet_skipped += 1
                    continue

                classification = get('classification')
                if classification not in CLASSIFICATION_VALUES:
                    sheet_errors.append(f'Sheet "{ws.title}" Row {row_num}: Invalid classification "{classification}" — skipped.')
                    sheet_skipped += 1
                    continue

                status = get('status') or 'Active'
                if status not in STATUS_VALUES:
                    status = 'Active'

                purchase_date = parse_date(get('purchase_date'), row, col_map)

                InventoryItem.objects.create(
                    id_number=get('id_number'),
                    asset_tag=get('asset_tag'),
                    computer_tag=get('computer_tag'),
                    display_tag=get('display_tag'),
                    brand=get('brand'),
                    model=get('model'),
                    category=category,
                    classification=classification,
                    type=get('type'),
                    status=status,
                    serial_number=get('serial_number'),
                    assigned_person=get('assigned_person'),
                    department=get('department'),
                    project_site=get('project_site'),
                    purchase_date=purchase_date,
                    po=get('po'),
                    work_order=get('work_order'),
                    inventory_receipt=get('inventory_receipt'),
                    inventory_issuance=get('inventory_issuance'),
                    reference_number=get('reference_number'),
                    remarks=get('remarks'),
                    import_batch=batch_id,
                )
                sheet_created += 1

            return sheet_created, sheet_skipped, sheet_errors

        import uuid
        from django.utils import timezone
        batch_id = f"{timezone.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
        batch_label = excel_file.name

        created = 0
        skipped = 0
        errors = []

        for ws, forced_category in sheets_to_process:
            c, s, e = import_sheet(ws, forced_category, batch_id=batch_id)
            created += c
            skipped += s
            errors += e

        if created:
            messages.success(request, f'✓ Successfully imported {created} item{"s" if created != 1 else ""} [Batch: {batch_id}].' +
                             (f' {skipped} row{"s" if skipped != 1 else ""} skipped.' if skipped else ''))
        elif skipped:
            messages.warning(request, f'No items imported. {skipped} row{"s" if skipped != 1 else ""} had errors.')
        else:
            messages.warning(request, 'No data rows found in the file.')

        for err in errors[:5]:
            messages.warning(request, err)
        if len(errors) > 5:
            messages.warning(request, f'... and {len(errors) - 5} more row errors.')

    except Exception as e:
        messages.error(request, f'Failed to process file: {str(e)}')

    return redirect(request.META.get('HTTP_REFERER', 'inventory_list'))


@login_required
def barcode_view(request, pk):
    """Return barcode image as base64 JSON for a single item."""
    item = get_object_or_404(InventoryItem, pk=pk)
    tag = item.asset_tag or item.id_number or f"ITEM-{item.pk}"
    b64 = generate_barcode_base64(tag)
    return JsonResponse({
        'barcode': b64,
        'value': tag,
        'item': {
            'id_number': item.id_number,
            'asset_tag': item.asset_tag,
            'brand': item.brand,
            'model': item.model,
            'assigned_person': item.assigned_person,
            'department': item.department,
            'status': item.status,
            'category': item.category,
            'classification': item.classification,
        }
    })


@login_required
def barcode_scan(request):
    """Search item by scanned barcode value."""
    value = request.GET.get('q', '').strip()
    if not value:
        return JsonResponse({'found': False, 'message': 'No scan value provided'})
    item = InventoryItem.objects.filter(
        Q(asset_tag=value) | Q(id_number=value) | Q(serial_number=value)
    ).first()
    if item:
        return JsonResponse({
            'found': True,
            'pk': item.pk,
            'id_number': item.id_number,
            'asset_tag': item.asset_tag,
            'brand': item.brand,
            'model': item.model,
            'assigned_person': item.assigned_person,
            'department': item.department,
            'status': item.status,
            'category': item.category,
            'classification': item.classification,
            'serial_number': item.serial_number,
            'type': item.type,
            'reference_number': item.reference_number,
        })
    return JsonResponse({'found': False, 'message': f'No item found for "{value}"'})


@login_required
def print_barcodes(request):
    """Render a printable page of barcodes for selected or all items."""
    pks = request.GET.get('pks', '')
    category = request.GET.get('category', '')
    if pks:
        pk_list = [p for p in pks.split(',') if p.strip()]
        items = InventoryItem.objects.filter(pk__in=pk_list)
    elif category:
        items = InventoryItem.objects.filter(category=category)
    else:
        items = InventoryItem.objects.all()

    items_with_barcodes = []
    for item in items:
        tag = item.asset_tag or item.id_number or f"ITEM-{item.pk}"
        b64 = generate_barcode_base64(tag)
        items_with_barcodes.append({'item': item, 'barcode': b64, 'tag': tag})

    return render(request, 'inventory/print_barcodes.html', {
        'items_with_barcodes': items_with_barcodes,
        'category': category,
    })


@login_required
def export_excel(request):
    """Export current inventory (with active filters) to Excel."""
    items = InventoryItem.objects.all()

    category = request.GET.get('category', '')
    classification = request.GET.get('classification', '')
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    dept_filter = request.GET.get('department', '')

    if category:
        items = items.filter(category=category)
    if classification:
        items = items.filter(classification=classification)
    if search:
        items = items.filter(
            Q(id_number__icontains=search) | Q(asset_tag__icontains=search) |
            Q(brand__icontains=search) | Q(model__icontains=search) |
            Q(assigned_person__icontains=search) | Q(department__icontains=search) |
            Q(serial_number__icontains=search) | Q(reference_number__icontains=search)
        )
    if status_filter:
        items = items.filter(status=status_filter)
    if dept_filter:
        items = items.filter(department__icontains=dept_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory Export'

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill('solid', start_color='1A2B4A')
    header_font = Font(bold=True, color='FFFFFF', name='Segoe UI', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    LABELS = [
        'ID Number', 'Asset Tag', 'Computer Tag', 'Display Tag',
        'Brand', 'Model', 'Category', 'Classification', 'Type',
        'Status', 'Serial Number', 'Assigned Person', 'Department',
        'Project/Site', 'Purchase Date', 'PO', 'Work Order',
        'Inventory Receipt', 'Inventory Issuance', 'Reference Number', 'Remarks',
        'Created At', 'Updated At',
    ]

    for col_idx, label in enumerate(LABELS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    STATUS_COLORS = {
        'Active': 'D1FAE5', 'Inactive': 'FEE2E2',
        'Under Repair': 'FEF3C7', 'For Repair': 'DBEAFE', 'Disposed': 'F3F4F6',
    }

    for row_idx, item in enumerate(items, 2):
        row_data = [
            item.id_number, item.asset_tag, item.computer_tag, item.display_tag,
            item.brand, item.model, item.category, item.classification, item.type,
            item.status, item.serial_number, item.assigned_person, item.department,
            item.project_site,
            item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '',
            item.po, item.work_order, item.inventory_receipt, item.inventory_issuance,
            item.reference_number, item.remarks,
            item.created_at.strftime('%Y-%m-%d %H:%M') if item.created_at else '',
            item.updated_at.strftime('%Y-%m-%d %H:%M') if item.updated_at else '',
        ]
        status_color = STATUS_COLORS.get(item.status, 'FFFFFF')
        row_fill = PatternFill('solid', start_color=status_color)
        row_font = Font(name='Segoe UI', size=9)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if col_idx == 10:  # Status column gets color
                cell.fill = row_fill

    widths = {1:12, 2:14, 3:14, 4:13, 5:14, 6:18, 7:16, 8:15, 9:14, 10:13,
              11:18, 12:22, 13:20, 14:18, 15:14, 16:12, 17:14, 18:18, 19:18,
              20:18, 21:30, 22:18, 23:18}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'A2'

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    from django.utils import timezone
    now_str = timezone.now().strftime('%Y-%m-%d %H:%M')

    summary_data = [
        ('Inventory Export Summary', True, '1A2B4A', 'FFFFFF'),
        (f'Generated: {now_str}', False, None, None),
        ('', False, None, None),
        ('Filter Applied', True, 'EFF6FF', '1E40AF'),
        (f'Category: {category or "All"}', False, None, None),
        (f'Status: {status_filter or "All"}', False, None, None),
        (f'Department: {dept_filter or "All"}', False, None, None),
        (f'Search: {search or "None"}', False, None, None),
        ('', False, None, None),
        ('Totals', True, 'EFF6FF', '1E40AF'),
        (f'Total Records: {items.count()}', False, None, None),
    ]
    status_counts = {}
    for item in items:
        status_counts[item.status] = status_counts.get(item.status, 0) + 1
    for status, count in status_counts.items():
        summary_data.append((f'{status}: {count}', False, None, None))

    for row_idx, (text, bold, bg, fg) in enumerate(summary_data, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, name='Segoe UI', size=11 if bold else 10,
                         color=fg or '374151')
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(vertical='center')
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'inventory_export_{timestamp}.xlsx'
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Import History & Batch Delete ───────────────────────────

@login_required
def import_history(request):
    """Show all import batches with item counts, allow bulk delete."""
    batches = (
        InventoryItem.objects.exclude(import_batch='')
        .values('import_batch')
        .annotate(
            count=Count('import_batch'),
            imported_at=models_Min('created_at'),
        )
        .order_by('-imported_at')
    )
    # Items added manually (no batch tag)
    manual_count = InventoryItem.objects.filter(import_batch='').count()
    context = {
        'batches': batches,
        'manual_count': manual_count,
        'total': InventoryItem.objects.count(),
    }
    return render(request, 'inventory/import_history.html', context)


@login_required
def delete_import_batch(request, batch_id):
    """Delete all items belonging to a specific import batch."""
    if request.method != 'POST':
        return redirect('import_history')
    deleted_count, _ = InventoryItem.objects.filter(import_batch=batch_id).delete()
    messages.success(request, f'✓ Deleted {deleted_count} item{"s" if deleted_count != 1 else ""} from batch "{batch_id}".')
    return redirect('import_history')


@login_required
def delete_all_items(request):
    """Delete every inventory item in the system."""
    if request.method != 'POST':
        return redirect('inventory_list')
    count, _ = InventoryItem.objects.all().delete()
    messages.success(request, f'✓ Deleted all {count} inventory item{"s" if count != 1 else ""} from the system.')
    return redirect('inventory_list')


@login_required
def import_export(request):
    """Dedicated Import/Export page with upload & download tabs."""
    from django.db.models import Count
    batches = (
        InventoryItem.objects.exclude(import_batch='')
        .values('import_batch')
        .annotate(
            count=Count('import_batch'),
            imported_at=models_Min('created_at'),
        )
        .order_by('-imported_at')
    )
    manual_count = InventoryItem.objects.filter(import_batch='').count()
    context = {
        'batches': batches,
        'manual_count': manual_count,
        'total': InventoryItem.objects.count(),
        'active_tab': request.GET.get('tab', 'upload'),
    }
    return render(request, 'inventory/import_export.html', context)


# ─── DOCS View ───────────────────────────────────────────────

@login_required
def docs_view(request):
    """
    DOCS page — three tabs: Inventory Receipt, Inventory Issuance, Work Order.
    Searchable by asset tag / ID number / assigned person.
    Data is pulled live from InventoryItem.
    """
    from django.db.models import Q

    tab = request.GET.get('tab', 'receipt')
    search = request.GET.get('search', '').strip()

    base_qs = InventoryItem.objects.all()
    if search:
        base_qs = base_qs.filter(
            Q(asset_tag__icontains=search) |
            Q(id_number__icontains=search) |
            Q(assigned_person__icontains=search) |
            Q(brand__icontains=search) |
            Q(model__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(department__icontains=search)
        )

    # Each tab filters to items that actually have the relevant doc number
    receipt_items   = base_qs.exclude(inventory_receipt='').order_by('inventory_receipt')
    issuance_items  = base_qs.exclude(inventory_issuance='').order_by('inventory_issuance')
    workorder_items = base_qs.exclude(work_order='').order_by('work_order')

    # Group by doc number for a cleaner display
    def group_by_doc(queryset, field):
        groups = {}
        for item in queryset:
            key = getattr(item, field) or '—'
            groups.setdefault(key, []).append(item)
        return groups

    receipt_groups   = group_by_doc(receipt_items,   'inventory_receipt')
    issuance_groups  = group_by_doc(issuance_items,  'inventory_issuance')
    workorder_groups = group_by_doc(workorder_items, 'work_order')

    context = {
        'tab': tab,
        'search': search,
        'receipt_items':    receipt_items,
        'issuance_items':   issuance_items,
        'workorder_items':  workorder_items,
        'receipt_groups':   receipt_groups,
        'issuance_groups':  issuance_groups,
        'workorder_groups': workorder_groups,
        'receipt_count':    receipt_items.count(),
        'issuance_count':   issuance_items.count(),
        'workorder_count':  workorder_items.count(),
    }
    return render(request, 'inventory/docs.html', context)


@login_required
def print_doc_view(request):
    """
    Print view for DOCS — Inventory Receipt, Inventory Issuance, Work Order.
    URL params: ?doc_type=ir|ii|wo&doc_ref=XXXXX (optional: single doc)
                &search=... (optional: filter by asset/person)
    """
    from django.db.models import Q

    doc_type = request.GET.get('doc_type', 'ir').lower()
    doc_ref  = request.GET.get('doc_ref', '').strip()
    search   = request.GET.get('search', '').strip()

    base_qs = InventoryItem.objects.all()
    if search:
        base_qs = base_qs.filter(
            Q(asset_tag__icontains=search) |
            Q(id_number__icontains=search) |
            Q(assigned_person__icontains=search) |
            Q(brand__icontains=search) |
            Q(model__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(department__icontains=search)
        )

    def group_by_doc(qs, field):
        groups = {}
        for item in qs:
            key = getattr(item, field) or '—'
            groups.setdefault(key, []).append(item)
        return groups

    if doc_type == 'ir':
        qs = base_qs.exclude(inventory_receipt='').order_by('inventory_receipt')
        if doc_ref:
            qs = qs.filter(inventory_receipt=doc_ref)
        groups = group_by_doc(qs, 'inventory_receipt')
        template = 'inventory/print_ir.html'

    elif doc_type == 'ii':
        qs = base_qs.exclude(inventory_issuance='').order_by('inventory_issuance')
        if doc_ref:
            qs = qs.filter(inventory_issuance=doc_ref)
        groups = group_by_doc(qs, 'inventory_issuance')
        template = 'inventory/print_ii.html'

    else:  # wo
        qs = base_qs.exclude(work_order='').order_by('work_order')
        if doc_ref:
            qs = qs.filter(work_order=doc_ref)
        groups = group_by_doc(qs, 'work_order')
        template = 'inventory/print_wo.html'

    context = {
        'doc_type': doc_type,
        'doc_ref':  doc_ref,
        'search':   search,
        'groups':   groups,
    }
    return render(request, template, context)


@login_required
def save_doc_item(request):
    """
    AJAX endpoint — create or update an InventoryItem from the DOCS modal form.
    POST params: doc_type (ir|ii|wo), and all item fields.
    Returns JSON {success, id, doc_ref, message}.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)

    doc_type = request.POST.get('doc_type', 'ir')
    item_id  = request.POST.get('item_id', '').strip()

    if item_id:
        item = get_object_or_404(InventoryItem, pk=item_id)
        form = InventoryItemForm(request.POST, instance=item)
    else:
        form = InventoryItemForm(request.POST)

    if form.is_valid():
        item = form.save()
        doc_ref = {
            'ir': item.inventory_receipt,
            'ii': item.inventory_issuance,
            'wo': item.work_order,
        }.get(doc_type, '')
        return JsonResponse({
            'success': True,
            'id': item.pk,
            'doc_ref': doc_ref,
            'message': 'Item saved successfully.',
        })
    else:
        return JsonResponse({
            'success': False,
            'errors': {k: [str(e) for e in v] for k, v in form.errors.items()},
            'message': 'Please fix the errors below.',
        })
